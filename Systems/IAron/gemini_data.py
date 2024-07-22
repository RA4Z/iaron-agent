from openpyxl import load_workbook
import json

arquivo_excel = 'database/Dev Assistant.xlsx'
workbook = load_workbook(arquivo_excel)
planilha = workbook.worksheets[0]
data = []
data_inputs = []

for linha in planilha.iter_rows(min_row=2, values_only=True):
    data.append(f"input: {linha[0]}")
    data.append(f"output: {linha[1]}")

inputs = json.load(open('database/inputs_examples.json', 'r', encoding='utf-8'))
for value in inputs:
    data_inputs.append(f"input: {value['input']}")
    data_inputs.append(f"output: {value['output']}")
